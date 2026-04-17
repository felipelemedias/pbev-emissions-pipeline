"""Gera dicionário de variáveis do PBEV 2026 em Excel."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Cores ──────────────────────────────────────────────────────────────────
COR_HEADER   = "1F4E79"   # azul escuro
COR_SECAO    = "2E75B6"   # azul médio
COR_DESTAQUE = "D6E4F0"   # azul claro
COR_BRANCO   = "FFFFFF"
COR_CINZA    = "F2F2F2"

def cor_fill(hex_cor):
    return PatternFill("solid", fgColor=hex_cor)

def borda():
    lado = Side(style="thin", color="BFBFBF")
    return Border(left=lado, right=lado, top=lado, bottom=lado)

def cabecalho(cell, texto, tamanho=11, negrito=True, cor_fonte="FFFFFF", cor_fundo=COR_HEADER):
    cell.value = texto
    cell.font = Font(bold=negrito, size=tamanho, color=cor_fonte, name="Calibri")
    cell.fill = cor_fill(cor_fundo)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = borda()

def celula(cell, texto, negrito=False, cor_fundo=COR_BRANCO, wrap=True, alinhamento="left"):
    cell.value = texto
    cell.font = Font(bold=negrito, size=10, name="Calibri")
    cell.fill = cor_fill(cor_fundo)
    cell.alignment = Alignment(horizontal=alinhamento, vertical="center", wrap_text=wrap)
    cell.border = borda()


# ════════════════════════════════════════════════════════════════════════════
# ABA 1 — DICIONÁRIO DE VARIÁVEIS
# ════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dicionário de Variáveis"

ws1.merge_cells("A1:G1")
cabecalho(ws1["A1"],
    "DICIONÁRIO DE VARIÁVEIS — Tabela PBEV 2026 (INMETRO)",
    tamanho=13, cor_fundo=COR_HEADER)
ws1.row_dimensions[1].height = 30

ws1.merge_cells("A2:G2")
celula(ws1["A2"],
    "Fonte: Programa Brasileiro de Etiquetagem Veicular (PBEV) — Tabela de Automóveis e "
    "Comerciais Leves, Janeiro/2026 — Rev.04. Instituto Nacional de Metrologia, Qualidade "
    "e Tecnologia (INMETRO).",
    cor_fundo=COR_DESTAQUE)

cabecalhos_col = ["Coluna (CSV)", "Nome Original (PDF)", "Descrição",
                  "Unidade", "Tipo", "Valores Possíveis / Legenda", "Observações"]
for j, h in enumerate(cabecalhos_col, 1):
    cabecalho(ws1.cell(3, j), h, tamanho=10, cor_fundo=COR_SECAO)

variaveis = [
    # (coluna_csv, nome_pdf, descricao, unidade, tipo, valores, obs)

    # ── Identificação ────────────────────────────────────────────────────
    ("categoria", "Categoria",
     "Segmento de mercado do veículo, conforme classificação INMETRO/ANFAVEA",
     "—", "Texto",
     "Sub Compacto · Compacto · Médio · Grande · Extra Grande · "
     "Utilitário Esportivo Compacto · Utilitário Esportivo Grande · "
     "Utilitário Esportivo Grande 4x4 · Fora de Estrada Compacto · "
     "Fora de Estrada Grande · Picape · Picape Compacta · "
     "Comercial · Minivan · Esportivo",
     "Classificação oficial utilizada pelo PBEV"),

    ("marca", "Marca",
     "Fabricante / montadora do veículo",
     "—", "Texto", "Ex.: FIAT, CHEVROLET, VW, TOYOTA, BYD…", ""),

    ("modelo", "Modelo",
     "Nome comercial do modelo do veículo",
     "—", "Texto", "Ex.: ONIX, HB20, POLO, SPIN…", ""),

    ("versao", "Versão",
     "Identificação da versão/trim level (acabamento, pacote de opcionais)",
     "—", "Texto", "Ex.: LX, EXL, PREMIER, HIGHLINE…",
     "19 registros sem versão — veículos com versão única ou não informada"),

    ("motor", "Motor",
     "Especificação do motor: cilindrada em litros e número de válvulas",
     "—", "Texto",
     "Formato: [Cilindrada]L-[Válvulas]V · Ex.: 1.0-12V, 1.8L-8V, 2.0-16V\n"
     "Elétrico = propulsor elétrico sem motor a combustão",
     ""),

    ("tipo_propulsao", "Tipo de Propulsão",
     "Tecnologia de propulsão do veículo",
     "—", "Texto",
     "Combustão = Motor a combustão interna exclusivo\n"
     "Elétrico = Propulsão elétrica pura (BEV)\n"
     "Híbrido = Híbrido convencional (HEV) — sem recarga externa\n"
     "Plug-in = Híbrido plug-in (PHEV) — bateria recarregável externamente",
     ""),

    ("transmissao", "Transmissão",
     "Tipo e número de velocidades da transmissão",
     "—", "Texto",
     "M-n = Manual n velocidades (ex.: M-5 = Manual 5 velocidades)\n"
     "A-n = Automática n velocidades (ex.: A-6 = Automática 6 marchas)\n"
     "CVT ou CVT-n = Variação Contínua\n"
     "DCT-n = Dupla Embreagem (ex.: DCT-7)\n"
     "MTA = Automatizada (Manual robotizada)\n"
     "DHT = Dedicated Hybrid Transmission (transmissão híbrida dedicada)\n"
     "N.A. = Não Aplicável (veículos elétricos puros)",
     ""),

    ("combustivel", "Combustível",
     "Combustível principal utilizado pelo veículo",
     "—", "Texto",
     "E = Etanol\n"
     "G = Gasolina\n"
     "F = Flex (aceita Etanol E Gasolina em qualquer proporção)\n"
     "D = Diesel",
     "Veículos elétricos puros têm combustivel=E (energia elétrica)"),

    # ── Emissões ────────────────────────────────────────────────────────
    ("nmog_nox_mg_km", "NMOG+NOx (mg/km)",
     "Soma das emissões de hidrocarbonetos não-metano orgânicos (NMOG) e "
     "óxidos de nitrogênio (NOx) medidos no ciclo de teste brasileiro",
     "mg/km", "Numérico (Real)",
     "Limite PROCONVE L8: ≤ 60 mg/km\n"
     "0 = veículo elétrico puro\n"
     "ND / NaN = Não detectado ou não aplicável",
     "71 registros nulos (EVs e PHEVs sem medição de escapamento)"),

    ("co_mg_km", "CO (mg/km)",
     "Emissões de monóxido de carbono medidas no ciclo de teste brasileiro",
     "mg/km", "Numérico (Real)",
     "Limite PROCONVE L8: ≤ 700 mg/km\n"
     "0 = veículo elétrico puro\n"
     "ND / NaN = Não detectado ou não aplicável",
     "45 registros nulos (EVs). SPIN 1.8L: 800 mg/km → acima do limite L8"),

    ("co2_fossil_g_km", "CO₂ Fóssil (g/km)",
     "Emissões de dióxido de carbono de origem fóssil. Para veículos Flex, "
     "representa o CO₂ fóssil medido no modo etanol (combustão de origem não-fóssil "
     "não é contabilizada). Para veículos a Gasolina, representa o CO₂ total.",
     "g/km", "Numérico (Real)",
     "Quanto menor, menor o impacto climático do veículo\n"
     "0 ou NaN = veículo elétrico puro",
     "158 registros nulos — principalmente EVs e PHEVs em modo elétrico"),

    ("consumo_mj_km", "Consumo Energético (MJ/km)",
     "Consumo de energia total do veículo por quilômetro percorrido, "
     "independente da fonte de energia. Indicador de eficiência energética global.",
     "MJ/km", "Numérico (Real)",
     "Quanto menor, mais eficiente o veículo\n"
     "Base do cálculo da Classificação PBE",
     "Única coluna sem valores nulos — todos os veículos possuem este valor"),

    # ── Classificações ───────────────────────────────────────────────────
    ("nota_verde_l8", "Nota Verde (PROCONVE L8)",
     "Classificação de desempenho ambiental em relação aos limites de emissão "
     "de poluentes do PROCONVE Fase L8 (vigente desde jan/2025). "
     "Avalia CO, NMOG+NOx e NOx conjuntamente.",
     "—", "Texto (Categoria)",
     "A = Melhor desempenho ambiental (emissões muito abaixo dos limites)\n"
     "B = Bom desempenho (emissões abaixo dos limites com margem)\n"
     "C = Desempenho regular (emissões próximas aos limites)\n"
     "D = Desempenho insuficiente (emissões acima de algum limite)\n"
     "E = Pior desempenho (emissões significativamente acima dos limites)",
     "Veículos elétricos puros recebem nota A automaticamente"),

    ("classificacao_pbe", "Classificação PBE",
     "Classificação de eficiência energética do Programa Brasileiro de "
     "Etiquetagem Veicular (PBE/INMETRO), baseada no consumo energético (MJ/km). "
     "Compara o veículo com outros da mesma categoria e faixa de consumo.",
     "—", "Texto (Categoria)",
     "A = Mais eficiente (menor consumo energético na categoria)\n"
     "B = Eficiente\n"
     "C = Eficiência média\n"
     "D = Pouco eficiente\n"
     "E = Menos eficiente (maior consumo energético na categoria)",
     "Independente da Nota Verde — um veículo pode ter PBE=A e Nota Verde=C"),

    # ── Colunas calculadas (não originam do PDF) ────────────────────────
    ("co_excesso_pct", "— (calculado)",
     "Percentual de excesso de CO em relação ao limite L8 de 700 mg/km. "
     "Calculado como: max(0, (co_mg_km − 700) / 700 × 100)",
     "%", "Numérico (Real)",
     "0.0 = dentro do limite\n"
     "> 0 = acima do limite (valor indica o excesso percentual)\n"
     "NaN = CO não medido (EV)",
     "Coluna gerada por extract_pbev.py — não consta no PDF original"),

    ("nmog_excesso_pct", "— (calculado)",
     "Percentual de excesso de NMOG+NOx em relação ao limite L8 de 60 mg/km. "
     "Calculado como: max(0, (nmog_nox_mg_km − 60) / 60 × 100)",
     "%", "Numérico (Real)",
     "0.0 = dentro do limite\n"
     "> 0 = acima do limite\n"
     "NaN = não medido (EV)",
     "Coluna gerada por extract_pbev.py — não consta no PDF original"),

    ("viola_l8", "— (calculado)",
     "Indica se o veículo viola ao menos um dos limites do PROCONVE L8 "
     "(CO > 700 mg/km OU NMOG+NOx > 60 mg/km)",
     "—", "Booleano",
     "True = viola pelo menos um limite L8\n"
     "False = dentro de todos os limites avaliados",
     "Coluna gerada por extract_pbev.py — não consta no PDF original"),
]

for i, row in enumerate(variaveis):
    linha = 4 + i
    cor = COR_CINZA if i % 2 == 0 else COR_BRANCO
    celula(ws1.cell(linha, 1), row[0], negrito=True, cor_fundo=cor)
    for j, val in enumerate(row[1:], 2):
        celula(ws1.cell(linha, j), val, cor_fundo=cor)
    ws1.row_dimensions[linha].height = 60

# Larguras
larguras = [20, 22, 45, 12, 14, 60, 40]
for j, w in enumerate(larguras, 1):
    ws1.column_dimensions[get_column_letter(j)].width = w


# ════════════════════════════════════════════════════════════════════════════
# ABA 2 — LEGENDAS DO PDF
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Legendas do PDF")

ws2.merge_cells("A1:C1")
cabecalho(ws2["A1"], "LEGENDAS E SÍMBOLOS — Tabela PBEV 2026", tamanho=13, cor_fundo=COR_HEADER)
ws2.row_dimensions[1].height = 30

secoes = [
    ("TIPO DE PROPULSÃO", [
        ("Combustão",   "Motor a combustão interna exclusivo (gasolina, etanol, flex ou diesel)"),
        ("Elétrico",    "Propulsão elétrica pura (BEV — Battery Electric Vehicle)"),
        ("Híbrido",     "Híbrido convencional (HEV) — motor elétrico auxiliar, sem recarga externa"),
        ("Plug-in",     "Híbrido plug-in (PHEV) — bateria de maior capacidade, recarga externa possível"),
    ]),
    ("TRANSMISSÃO", [
        ("M-n",   "Manual com n velocidades (ex.: M-5 = 5 marchas manual)"),
        ("A-n",   "Automática com n velocidades (ex.: A-6 = 6 marchas automática)"),
        ("CVT",   "Transmissão de variação contínua (Continuously Variable Transmission)"),
        ("CVT-n", "CVT com modo sequencial de n velocidades simuladas"),
        ("DCT-n", "Dupla embreagem com n velocidades (Dual Clutch Transmission)"),
        ("MTA",   "Manual automatizada / robotizada (Automated Manual Transmission)"),
        ("DHT",   "Dedicated Hybrid Transmission — transmissão específica para híbridos"),
        ("N.A.",  "Não Aplicável — veículos elétricos puros sem transmissão convencional"),
    ]),
    ("AR CONDICIONADO (Ar Cond.)", [
        ("S", "Sim — veículo possui ar-condicionado de série"),
        ("N", "Não — veículo não possui ar-condicionado de série"),
    ]),
    ("DIREÇÃO ASSISTIDA", [
        ("H",   "Hidráulica"),
        ("M",   "Mecânica (sem assistência)"),
        ("E",   "Elétrica"),
        ("E-H", "Eletro-hidráulica"),
    ]),
    ("COMBUSTÍVEL", [
        ("E", "Etanol (inclui veículos elétricos — energia elétrica)"),
        ("G", "Gasolina"),
        ("F", "Flex — aceita etanol e gasolina em qualquer proporção"),
        ("D", "Diesel"),
    ]),
    ("NOTA VERDE — PROCONVE L8", [
        ("A", "Melhor desempenho ambiental — emissões muito abaixo dos limites L8"),
        ("B", "Bom desempenho — emissões abaixo dos limites com boa margem"),
        ("C", "Desempenho regular — emissões próximas aos limites L8"),
        ("D", "Desempenho insuficiente — emissões acima de algum limite L8"),
        ("E", "Pior desempenho — emissões significativamente acima dos limites L8"),
    ]),
    ("CLASSIFICAÇÃO PBE (Eficiência Energética)", [
        ("A", "Mais eficiente energeticamente na categoria (menor MJ/km)"),
        ("B", "Eficiente"),
        ("C", "Eficiência média"),
        ("D", "Pouco eficiente"),
        ("E", "Menos eficiente na categoria (maior MJ/km)"),
    ]),
    ("SÍMBOLOS ESPECIAIS", [
        ("\\",  "Não aplicável — dado não existe para esse veículo/combustível"),
        ("ND",  "Não Detectado — emissão abaixo do limite de detecção do equipamento"),
        ("0",   "Zero — veículo elétrico puro, sem emissões diretas no escapamento"),
        ("SIM", "Selo CONPET — veículo recebeu o Selo de Eficiência Energética CONPET (top 12,9% da categoria)"),
        ("-",   "Sem Selo CONPET"),
    ]),
    ("LIMITES PROCONVE L8 (vigente desde jan/2025)", [
        ("CO ≤ 700 mg/km",        "Monóxido de carbono — limite para veículos a gasolina/flex"),
        ("NMOG+NOx ≤ 60 mg/km",   "Hidrocarbonetos orgânicos não-metano + óxidos de nitrogênio"),
        ("NOx ≤ 0,06 g/km",       "Óxidos de nitrogênio isolados"),
        ("CO₂ fóssil — sem limite fixo", "Informativo — base para a Classificação PBE e Nota Verde"),
    ]),
]

linha = 3
for secao, itens in secoes:
    ws2.merge_cells(f"A{linha}:C{linha}")
    cabecalho(ws2.cell(linha, 1), secao, tamanho=10, cor_fundo=COR_SECAO)
    ws2.row_dimensions[linha].height = 20
    linha += 1

    for cod, desc in itens:
        cor = COR_CINZA if (linha % 2 == 0) else COR_BRANCO
        celula(ws2.cell(linha, 1), cod, negrito=True, cor_fundo=cor, alinhamento="center")
        ws2.merge_cells(f"B{linha}:C{linha}")
        celula(ws2.cell(linha, 2), desc, cor_fundo=cor)
        ws2.row_dimensions[linha].height = 18
        linha += 1

    linha += 1  # linha em branco entre seções

ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 60
ws2.column_dimensions["C"].width = 20

# ════════════════════════════════════════════════════════════════════════════
# ABA 3 — LIMITES L8 (referência rápida)
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Limites PROCONVE L8")

ws3.merge_cells("A1:E1")
cabecalho(ws3["A1"], "LIMITES DE EMISSÃO — PROCONVE FASE L8 (Resolução CONAMA + IBAMA)",
          tamanho=13, cor_fundo=COR_HEADER)
ws3.row_dimensions[1].height = 30

for j, h in enumerate(["Poluente", "Símbolo", "Limite L8", "Fase L7 (anterior)", "Redução"], 1):
    cabecalho(ws3.cell(2, j), h, tamanho=10, cor_fundo=COR_SECAO)

limites = [
    ("Monóxido de Carbono",               "CO",        "700 mg/km",    "1 300 mg/km", "−46%"),
    ("NMOG + Óxidos de Nitrogênio",       "NMOG+NOx",  "60 mg/km",     "80 mg/km",    "−25%"),
    ("Óxidos de Nitrogênio",              "NOx",       "0,06 g/km",    "0,08 g/km",   "−25%"),
    ("Hidrocarbonetos Totais",            "THC",       "60 mg/km",     "80 mg/km",    "−25%"),
    ("Material Particulado (flex/diesel)","MP",        "4,5 mg/km",    "6 mg/km",     "−25%"),
    ("Aldeídos (flex/etanol)",            "RCHO",      "20 mg/km",     "30 mg/km",    "−33%"),
]

for i, row in enumerate(limites):
    linha2 = 3 + i
    cor = COR_CINZA if i % 2 == 0 else COR_BRANCO
    for j, val in enumerate(row, 1):
        negrito = j <= 2
        celula(ws3.cell(linha2, j), val, negrito=negrito,
               cor_fundo=cor, alinhamento="center" if j >= 3 else "left")
    ws3.row_dimensions[linha2].height = 20

ws3.merge_cells("A10:E10")
celula(ws3["A10"],
    "Vigência: 1º de janeiro de 2025 para veículos novos fabricados no Brasil. "
    "Fonte: IBAMA — Programa de Controle da Poluição do Ar por Veículos Automotores (PROCONVE).",
    cor_fundo=COR_DESTAQUE)

for j, w in enumerate([35, 14, 16, 18, 12], 1):
    ws3.column_dimensions[get_column_letter(j)].width = w

# ────────────────────────────────────────────────────────────────────────────
OUT = "dicionario_pbev_2026.xlsx"
wb.save(OUT)
print(f"Salvo: {OUT}")
